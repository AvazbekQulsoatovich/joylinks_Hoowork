"""
Excel Export funksiyalari - Admin uchun hisobotlar
"""
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Avg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import Homework, Submission
from academy.models import Course, Group
from users.models import User


def get_styled_workbook():
    """Stillangan Excel workbook yaratish"""
    wb = Workbook()
    return wb


def style_header_row(ws, row_num=1):
    """Sarlavha qatorini stillash"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def auto_adjust_columns(ws):
    """Ustun kengliklarini avtomatik sozlash"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_all_submissions(course_id=None, group_id=None):
    """
    Barcha topshiriqlarni Excel formatda export qilish
    
    Ustunlar:
    - O'quvchi ismi
    - Kurs
    - Guruh
    - Uyga vazifa nomi
    - Olgan %
    - O'rtacha %
    - Status
    """
    wb = get_styled_workbook()
    ws = wb.active
    ws.title = "Natijalar"
    
    # Sarlavhalar
    headers = [
        "O'quvchi ismi",
        "Username",
        "Kurs",
        "Guruh",
        "Uyga vazifa nomi",
        "Olgan %",
        "Status",
        "Topshirilgan sana",
        "Baholangan sana"
    ]
    ws.append(headers)
    style_header_row(ws)
    
    # Ma'lumotlarni olish
    submissions = Submission.objects.select_related(
        'student', 'homework', 'homework__group', 'homework__group__course'
    ).order_by('homework__group__course__name', 'homework__group__name', 'student__last_name')
    
    if course_id:
        submissions = submissions.filter(homework__group__course_id=course_id)
    if group_id:
        submissions = submissions.filter(homework__group_id=group_id)
    
    for sub in submissions:
        student_name = f"{sub.student.last_name} {sub.student.first_name}".strip() or sub.student.username
        status = "Baholangan" if sub.is_graded else "Tekshirilmagan"
        if sub.is_graded and sub.score_percent == 0:
            status = "Topshirmagan (0%)"
        
        ws.append([
            student_name,
            sub.student.username,
            sub.homework.group.course.name,
            sub.homework.group.name,
            sub.homework.title,
            sub.score_percent if sub.is_graded else "-",
            status,
            sub.submitted_at.strftime("%d.%m.%Y %H:%M") if sub.submitted_at else "-",
            sub.graded_at.strftime("%d.%m.%Y %H:%M") if sub.graded_at else "-"
        ])
    
    auto_adjust_columns(ws)
    
    # O'rtachalar sahifasi
    ws2 = wb.create_sheet("O'rtachalar")
    headers2 = ["O'quvchi", "Username", "Kurs", "Guruh", "Jami vazifalar", "Topshirganlar", "O'rtacha %"]
    ws2.append(headers2)
    style_header_row(ws2)
    
    # O'rtachalarni hisoblash
    students = User.objects.filter(role='STUDENT')
    for student in students:
        for group in student.study_groups.all():
            subs = Submission.objects.filter(
                student=student, 
                homework__group=group,
                is_graded=True
            )
            total_hw = Homework.objects.filter(group=group).count()
            submitted = subs.count()
            avg = subs.aggregate(avg=Avg('score_percent'))['avg'] or 0
            
            if total_hw > 0:
                student_name = f"{student.last_name} {student.first_name}".strip() or student.username
                ws2.append([
                    student_name,
                    student.username,
                    group.course.name,
                    group.name,
                    total_hw,
                    submitted,
                    round(avg, 1)
                ])
    
    auto_adjust_columns(ws2)
    
    return wb


def export_group_report(group_id):
    """Guruh bo'yicha batafsil hisobot"""
    group = Group.objects.select_related('course').get(pk=group_id)
    wb = get_styled_workbook()
    ws = wb.active
    ws.title = f"{group.name}"
    
    # Ma'lumot
    ws.append([f"Guruh: {group.name}"])
    ws.append([f"Kurs: {group.course.name}"])
    ws.append([])
    
    # Sarlavhalar
    homeworks = Homework.objects.filter(group=group).order_by('sequence')
    headers = ["O'quvchi"] + [hw.title for hw in homeworks] + ["O'rtacha %"]
    ws.append(headers)
    style_header_row(ws, row_num=4)
    
    # O'quvchilar va baholar
    students = group.students.all().order_by('last_name')
    for student in students:
        row = [f"{student.last_name} {student.first_name}".strip() or student.username]
        scores = []
        for hw in homeworks:
            sub = Submission.objects.filter(homework=hw, student=student).first()
            if sub and sub.is_graded:
                row.append(sub.score_percent)
                scores.append(sub.score_percent)
            else:
                row.append("-")
        
        avg = sum(scores) / len(scores) if scores else 0
        row.append(round(avg, 1))
        ws.append(row)
    
    auto_adjust_columns(ws)
    
    return wb


def export_course_report(course_id):
    """Kurs bo'yicha umumiy hisobot"""
    course = Course.objects.get(pk=course_id)
    groups = Group.objects.filter(course=course)
    
    wb = get_styled_workbook()
    ws = wb.active
    ws.title = "Umumiy"
    
    ws.append([f"Kurs: {course.name}"])
    ws.append([])
    
    headers = ["Guruh", "Talabalar soni", "Vazifalar soni", "O'rtacha %"]
    ws.append(headers)
    style_header_row(ws, row_num=3)
    
    for group in groups:
        student_count = group.students.count()
        hw_count = Homework.objects.filter(group=group).count()
        
        subs = Submission.objects.filter(homework__group=group, is_graded=True)
        avg = subs.aggregate(avg=Avg('score_percent'))['avg'] or 0
        
        ws.append([group.name, student_count, hw_count, round(avg, 1)])
    
    auto_adjust_columns(ws)
    
    return wb


def workbook_to_response(wb, filename):
    """Workbook'ni HTTP response sifatida qaytarish"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
